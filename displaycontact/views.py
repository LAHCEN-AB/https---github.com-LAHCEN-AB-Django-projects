from django.shortcuts import render
from openpyxl import load_workbook
# Create your views here.
def disp(request):
    wb=load_workbook(r'C:\Projects\emp\names.xlsx')
    ws=wb['emp']
    nbrrows=int(ws['F1'].value)
    items=[]
    for i in range(2,nbrrows+1):
        item=[]
        item.append( ws['A'+str(i)].value)
        item.append( ws['B'+str(i)].value)
        item.append( ws['C'+str(i)].value)
        item.append( ws['D'+str(i)].value)
        items.append(item)
       
    return render(request,'displaydata.html',{'items':items})
    
    
    
  
# def regEmp(request):
#     m=''
#     form=RegEmpForm()
#     if request.method=='POST':
#         form=RegEmpForm(request.POST)
#         if form.is_valid():
#             cd=form.cleaned_data
           
            
#             i=str(int(ws['F1'].value)+1)
#             ws['A'+i]=cd['name']
#             ws['B'+i]=cd['salary']
#             ws['C'+i]=cd['emp_email']
#             ws['D'+i]=cd['adress']
#             ws['F1']=i
#             wb.save(r'C:\Projects\emp\names.xlsx')
#             m='File saved succefully'
#             form=RegEmpForm()
#     return render(request,'newemp.html',{'form':form,'output':m})